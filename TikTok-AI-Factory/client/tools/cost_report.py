"""
TikTok AI Factory Pro — Cost Report Generator
===============================================
Standalone cost analysis and Excel export.

Usage:
  python tools/cost_report.py --log batch_log.json --output cost_report.xlsx
  python tools/cost_report.py --estimate --products 10 --per-product 5
"""

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

# ================================================================
# Pricing constants (USD, as of 2026)
# ================================================================

PRICING = {
    "openai": {
        "GPT-4o (input)":  {"per_1k_tokens": 0.00250, "tokens_per_task": 2000},
        "GPT-4o (output)": {"per_1k_tokens": 0.01000, "tokens_per_task": 800},
        "DALL-E 3 (1024×1792)": {"per_image": 0.04, "images_per_task": 6},
    },
    "elevenlabs": {
        "TTS (per 1k chars)": {"per_1k_chars": 0.015, "chars_per_task": 250},
    },
    "seedance": {
        "Seedance 2.0 (per sec)": {"per_second": 0.005, "seconds_per_task": 15},
    },
}


def estimate_single_video_cost() -> Dict:
    """Estimate cost for ONE video."""
    breakdown = {}

    # OpenAI
    gpt_input_cost = PRICING["openai"]["GPT-4o (input)"]["per_1k_tokens"] * PRICING["openai"]["GPT-4o (input)"]["tokens_per_task"] / 1000
    gpt_output_cost = PRICING["openai"]["GPT-4o (output)"]["per_1k_tokens"] * PRICING["openai"]["GPT-4o (output)"]["tokens_per_task"] / 1000
    dalle_cost = PRICING["openai"]["DALL-E 3 (1024×1792)"]["per_image"] * PRICING["openai"]["DALL-E 3 (1024×1792)"]["images_per_task"]
    openai_total = gpt_input_cost + gpt_output_cost + dalle_cost
    breakdown["OpenAI GPT-4o (文本)"] = round(gpt_input_cost + gpt_output_cost, 4)
    breakdown["OpenAI DALL-E 3 (6张图)"] = round(dalle_cost, 4)

    # ElevenLabs
    tts_cost = PRICING["elevenlabs"]["TTS (per 1k chars)"]["per_1k_chars"] * PRICING["elevenlabs"]["TTS (per 1k chars)"]["chars_per_task"] / 1000
    breakdown["ElevenLabs TTS"] = round(tts_cost, 4)

    # Seedance
    seedance_cost = PRICING["seedance"]["Seedance 2.0 (per sec)"]["per_second"] * PRICING["seedance"]["Seedance 2.0 (per sec)"]["seconds_per_task"]
    breakdown["Seedance 视频生成"] = round(seedance_cost, 4)

    total = sum(breakdown.values())
    return {"breakdown": breakdown, "total_per_video": round(total, 4)}


def estimate_batch_cost(product_count: int, videos_per_product: int) -> Dict:
    """Estimate total batch cost."""
    single = estimate_single_video_cost()
    total_videos = product_count * videos_per_product
    total_cost = single["total_per_video"] * total_videos
    return {
        "products": product_count,
        "videos_per_product": videos_per_product,
        "total_videos": total_videos,
        "per_video_cost": single["total_per_video"],
        "per_video_breakdown": single["breakdown"],
        "total_cost": round(total_cost, 2),
        "total_cost_gpt_only": round(single["breakdown"]["OpenAI GPT-4o (文本)"] * total_videos, 2),
        "total_cost_image_only": round(single["breakdown"]["OpenAI DALL-E 3 (6张图)"] * total_videos, 2),
        "total_cost_tts_only": round(single["breakdown"]["ElevenLabs TTS"] * total_videos, 2),
        "total_cost_video_only": round(single["breakdown"]["Seedance 视频生成"] * total_videos, 2),
    }


def generate_report_text(product_count: int, videos_per_product: int) -> str:
    """Generate a human-readable cost report."""
    est = estimate_batch_cost(product_count, videos_per_product)

    lines = []
    lines.append("=" * 60)
    lines.append("  TikTok AI Factory Pro — 成本预估报告")
    lines.append("=" * 60)
    lines.append(f"  产品数量:    {est['products']}")
    lines.append(f"  每产品视频:  {est['videos_per_product']}")
    lines.append(f"  总视频数:    {est['total_videos']}")
    lines.append("")
    lines.append("  [每视频成本明细]")
    for name, cost in est["per_video_breakdown"].items():
        bar = "█" * int(cost * 200) if cost > 0 else ""
        lines.append(f"  {name:<30} ${cost:<8.4f} {bar}")
    lines.append(f"  {'─' * 50}")
    lines.append(f"  {'每视频合计':<30} ${est['per_video_cost']:<8.4f}")
    lines.append("")
    lines.append("  [批量总成本]")
    lines.append(f"  GPT-4o 文本:     ${est['total_cost_gpt_only']:<10.2f}")
    lines.append(f"  DALL-E 3 图像:   ${est['total_cost_image_only']:<10.2f}")
    lines.append(f"  ElevenLabs TTS:  ${est['total_cost_tts_only']:<10.2f}")
    lines.append(f"  Seedance 视频:   ${est['total_cost_video_only']:<10.2f}")
    lines.append(f"  {'─' * 50}")
    lines.append(f"  总计:            ${est['total_cost']:<10.2f}")
    lines.append("")
    lines.append("  [价格参考]")
    lines.append(f"  单视频成本约 ${est['per_video_cost']}（含全部 AI 服务）")
    lines.append(f"  批量后可优化 Token 缓存降低文本成本 50%")
    lines.append("=" * 60)

    return "\n".join(lines)


def export_to_excel(estimates: Dict, output_path: Path):
    """Export cost estimate to XLSX."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        # Fallback to CSV
        csv_path = output_path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["项目", "金额(USD)"])
            for name, cost in estimates.get("per_video_breakdown", {}).items():
                writer.writerow([name, cost])
            writer.writerow(["总计", estimates.get("total_cost", 0)])
        print(f"已导出 CSV: {csv_path}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "成本预估"

    # Styles
    hdr_font = Font(bold=True, color="FFFFFF", size=13)
    hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    accent_fill = PatternFill(start_color="e94560", end_color="e94560", fill_type="solid")
    money_fmt = '$#,##0.0000'

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"].value = "TikTok AI Factory Pro — 成本预估报告"
    ws["A1"].font = Font(bold=True, size=16, color="e94560")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    ws["A2"].value = f"产品{estimates['products']}个 × 每产品{estimates['videos_per_product']}视频 = {estimates['total_videos']}视频"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Per-video breakdown
    ws["A4"] = "每视频成本明细"
    ws["A4"].font = Font(bold=True, size=12)
    for i, (name, cost) in enumerate(estimates.get("per_video_breakdown", {}).items(), 5):
        ws.cell(row=i, column=1, value=name)
        c = ws.cell(row=i, column=2, value=cost)
        c.number_format = money_fmt

    # Batch totals
    row = 12
    ws.cell(row=row, column=1, value="批量总成本").font = Font(bold=True, size=13, color="e94560")
    row += 1
    totals = [
        ("GPT-4o 文本", estimates.get("total_cost_gpt_only", 0)),
        ("DALL-E 3 图像", estimates.get("total_cost_image_only", 0)),
        ("ElevenLabs TTS", estimates.get("total_cost_tts_only", 0)),
        ("Seedance 视频", estimates.get("total_cost_video_only", 0)),
    ]
    for name, val in totals:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=val).number_format = '$#,##0.00'
        row += 1
    row += 1
    total_cell = ws.cell(row=row, column=1, value="总计")
    total_cell.font = Font(bold=True, size=14, color="e94560")
    val_cell = ws.cell(row=row, column=2, value=estimates.get("total_cost", 0))
    val_cell.font = Font(bold=True, size=14, color="e94560")
    val_cell.number_format = '$#,##0.00'

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"已导出 XLSX: {output_path}")


def export_from_batch_log(log_path: Path, output_path: Path):
    """Generate cost report from an existing batch_log.json."""
    data = json.loads(Path(log_path).read_text(encoding="utf-8"))
    tasks = data.get("tasks", [])

    completed = [t for t in tasks if t.get("status") == "completed"]
    failed = [t for t in tasks if t.get("status") == "failed"]
    single = estimate_single_video_cost()
    per_video = single["total_per_video"]

    rows = []
    for t in tasks:
        rows.append({
            "task_id": t.get("task_id", ""),
            "product": Path(t.get("product_path", "")).stem,
            "character": Path(t.get("character_path", "")).stem,
            "reference": Path(t.get("reference_video_path", "")).stem,
            "status": t.get("status", "unknown"),
            "attempts": t.get("attempts", 0),
            "cost": per_video if t.get("status") == "completed" else 0,
        })

    total_cost = sum(r["cost"] for r in rows)

    # Export to XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "批量成本报告"

        ws["A1"] = "TikTok AI Factory Pro — 批量生产成本报告"
        ws["A1"].font = Font(bold=True, size=14, color="e94560")
        ws.merge_cells("A1:G1")
        ws["A2"] = f"完成: {len(completed)} | 失败: {len(failed)} | 总成本: ${total_cost:.2f}"
        ws.merge_cells("A2:G2")

        headers = ["任务ID", "产品", "人物", "参考视频", "状态", "重试次数", "成本($)"]
        for col, h in enumerate(headers, 4):
            c = ws.cell(row=4, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="e94560", end_color="e94560", fill_type="solid")

        green = PatternFill(start_color="238636", end_color="238636", fill_type="solid")
        red = PatternFill(start_color="da3633", end_color="da3633", fill_type="solid")

        for i, row_data in enumerate(rows, 5):
            vals = [row_data[k] for k in ["task_id", "product", "character", "reference", "status", "attempts", "cost"]]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                if col == 5:
                    cell.fill = green if val == "completed" else red if val == "failed" else None
                    cell.font = Font(color="FFFFFF", bold=True)

        ws.column_dimensions["A"].width = 14
        for c in "BCDEF":
            ws.column_dimensions[c].width = 18
        ws.column_dimensions["G"].width = 12

        wb.save(output_path)
        print(f"成本报告已导出: {output_path}")
    except ImportError:
        csv_path = output_path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=["task_id", "product", "character", "reference", "status", "attempts", "cost"])
            writer.writeheader()
            writer.writerows(rows)
        print(f"成本报告已导出: {csv_path}")


# ================================================================
# CLI
# ================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="TikTok AI Factory Pro — 成本报告")
    parser.add_argument("--estimate", action="store_true", help="预估成本（不基于实际日志）")
    parser.add_argument("--products", type=int, default=10, help="产品数量")
    parser.add_argument("--per-product", type=int, default=5, help="每产品视频数")
    parser.add_argument("--log", type=str, help="batch_log.json 路径")
    parser.add_argument("--output", type=str, default="cost_report.xlsx", help="输出路径")

    args = parser.parse_args()

    if args.log:
        export_from_batch_log(Path(args.log), Path(args.output))
    elif args.estimate:
        est = estimate_batch_cost(args.products, args.per_product)
        print(generate_report_text(args.products, args.per_product))
        export_to_excel(est, Path(args.output))
    else:
        # Print default estimate
        print(generate_report_text(10, 5))
