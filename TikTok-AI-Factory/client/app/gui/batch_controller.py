"""
TikTok AI Factory Pro — Batch Controller
==========================================
Orchestrates the batch video factory — scan inputs, build task matrix,
run tasks via scheduler, track costs, and export reports.

Reuses OneClickController for individual task execution.
"""

import json
import logging
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable, List, Dict, Any

logger = logging.getLogger(__name__)


class BatchController:
    """Master controller for batch video production."""

    COUNTRIES_MAP = {
        "美国": ("US", "en", "american"),
        "英国": ("GB", "en", "british"),
        "马来西亚": ("MY", "ms", "malay"),
        "印尼": ("ID", "id", "indonesian"),
        "德国": ("DE", "de", "german"),
        "法国": ("FR", "fr", "french"),
        "西班牙": ("ES", "es", "spanish"),
    }

    STYLES = [
        "TikTok UGC", "Beauty Review", "Problem Solution",
        "Before After", "Testimonial", "POV Story",
    ]

    MODES = {
        "随机模式": "random",
        "固定模式": "fixed",
        "爆款复刻模式": "viral_clone",
    }

    def __init__(self, project_root: Path = None):
        self.project_root = project_root or Path(__file__).resolve().parent.parent.parent
        self.input_dir = self.project_root / "input"
        self.output_dir = self.project_root / "output"
        self.scheduler = None
        self._cost_log = []

        # Ensure directories
        for d in ["products", "characters", "reference_videos"]:
            (self.input_dir / d).mkdir(parents=True, exist_ok=True)

    # ================================================================
    # Input Scanning
    # ================================================================

    def scan_inputs(self) -> Dict[str, List[str]]:
        """Scan input directories and return file lists."""
        products = sorted([
            str(p) for p in (self.input_dir / "products").glob("*")
            if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
        ])
        characters = sorted([
            str(p) for p in (self.input_dir / "characters").glob("*")
            if p.suffix.lower() in (".jpg", ".jpeg", ".png")
        ])
        ref_videos = sorted([
            str(p) for p in (self.input_dir / "reference_videos").glob("*")
            if p.suffix.lower() in (".mp4", ".mov")
        ])
        return {"products": products, "characters": characters, "reference_videos": ref_videos}

    # ================================================================
    # Task Execution
    # ================================================================

    def run_batch(
        self,
        country: str = "美国",
        style: str = "TikTok UGC",
        videos_per_product: int = 3,
        mode: str = "random",
        max_concurrent: int = 3,
        progress_callback: Callable = None,
        stats_callback: Callable = None,
        log_callback: Callable = None,
        task_callback: Callable = None,
        done_callback: Callable = None,
    ) -> Dict[str, Any]:
        """
        Start batch production.

        Args:
            country: Target country
            style: Video style
            videos_per_product: Number of videos per product
            mode: "random" | "fixed" | "viral_clone"
            max_concurrent: Max parallel tasks
            progress_callback(stats_dict): Real-time stats
            stats_callback(stats_dict): Alias for progress
            log_callback(message): Log messages
            task_callback(event, task_id, data): Task lifecycle events
            done_callback(results_dict): Called when all done

        Returns:
            dict with batch results summary
        """
        inputs = self.scan_inputs()
        products = inputs["products"]
        characters = inputs["characters"]
        reference_videos = inputs["reference_videos"]

        if not products:
            return {"status": "error", "message": "未找到产品图片，请上传到 input/products/"}
        if not characters:
            return {"status": "error", "message": "未找到人物图片，请上传到 input/characters/"}
        if not reference_videos:
            return {"status": "error", "message": "未找到参考视频，请上传到 input/reference_videos/"}

        mode_key = self.MODES.get(mode, "random")

        # Build tasks
        from app.gui.batch_scheduler import BatchScheduler, BatchTask

        self.scheduler = BatchScheduler(max_concurrent=max_concurrent)
        self.scheduler.LOG_PATH = str(self.output_dir / "batch_log.json")

        tasks = self.scheduler.build_tasks(
            products=products,
            characters=characters,
            reference_videos=reference_videos,
            country=country,
            style=style,
            videos_per_product=videos_per_product,
            mode=mode_key,
        )
        self.scheduler.tasks = tasks

        # Set callbacks
        self.scheduler.on_stats_update = lambda s: (
            progress_callback(s) if progress_callback else None,
            stats_callback(s) if stats_callback else None,
        )
        self.scheduler.on_log = log_callback
        self.scheduler.on_task_start = lambda tid: (
            task_callback("start", tid, {}) if task_callback else None
        )
        self.scheduler.on_task_complete = lambda tid, out, vid: (
            task_callback("complete", tid, {"output": out, "video": vid}) if task_callback else None
        )
        self.scheduler.on_task_fail = lambda tid, err: (
            task_callback("fail", tid, {"error": err}) if task_callback else None
        )
        self.scheduler.on_all_complete = lambda s: (
            done_callback(self._build_results()) if done_callback else None
        )

        # Cost estimates
        estimates = BatchScheduler.estimate_costs(len(tasks), videos_per_product)

        # Start
        def worker(task: BatchTask) -> dict:
            return self._execute_single_task(task)

        self.scheduler.start(worker_fn=worker)

        return {
            "status": "started",
            "task_count": len(tasks),
            "products": len(products),
            "characters": len(characters),
            "reference_videos": len(reference_videos),
            "estimated_videos": len(tasks),
            "estimates": estimates,
        }

    def cancel(self):
        if self.scheduler:
            self.scheduler.cancel()

    # ================================================================
    # Single Task Worker
    # ================================================================

    def _execute_single_task(self, task) -> dict:
        """Execute one batch task using OneClickController."""
        from app.gui.one_click_controller import OneClickController
        ctrl = OneClickController(self.project_root)
        result = ctrl.run(
            product_path=task.product_path,
            character_path=task.character_path,
            video_path=task.reference_video_path,
            country=task.country,
            video_count=1,
            style=task.style,
        )
        # Track cost
        self._log_cost(task, result)
        return result

    def _log_cost(self, task, result):
        """Log estimated cost for a completed task."""
        entry = {
            "task_id": task.task_id,
            "product": Path(task.product_path).stem,
            "character": Path(task.character_path).stem,
            "reference": Path(task.reference_video_path).stem,
            "status": result.get("status", "unknown"),
            "elapsed_seconds": result.get("elapsed_seconds", 0),
            "estimated_cost": 0.42,  # per video average
            "timestamp": datetime.now().isoformat(),
        }
        self._cost_log.append(entry)

    # ================================================================
    # Results
    # ================================================================

    def _build_results(self) -> dict:
        if not self.scheduler:
            return {"status": "no_data"}
        stats = self.scheduler.stats.to_dict()
        return {
            "status": "completed",
            "stats": stats,
            "output_dir": str(self.output_dir),
            "cost_log": self._cost_log,
            "total_estimated_cost": round(len(self._cost_log) * 0.42, 2),
        }

    def get_cost_report(self) -> List[Dict]:
        return self._cost_log

    def export_cost_report(self, output_path: Path = None) -> Path:
        """Export cost report as CSV (and XLSX if openpyxl available)."""
        if output_path is None:
            output_path = self.output_dir / "cost_report.csv"
        import csv
        try:
            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                if self._cost_log:
                    fieldnames = self._cost_log[0].keys()
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(self._cost_log)
            # Also try XLSX
            try:
                self._export_xlsx(output_path.with_suffix(".xlsx"))
            except ImportError:
                pass
            return output_path
        except Exception as e:
            logger.error(f"导出成本报告失败: {e}")
            return output_path

    def _export_xlsx(self, path: Path):
        """Export to XLSX using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "成本报告"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="e94560", end_color="e94560", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Title
        ws.merge_cells("A1:H1")
        title = ws["A1"]
        title.value = "TikTok AI Factory Pro — 批量生产成本报告"
        title.font = Font(bold=True, size=16, color="e94560")
        title.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:H2")
        ws["A2"].value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["任务ID", "产品", "人物", "参考视频", "状态", "耗时(秒)", "预估成本($)", "时间"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        for row_idx, entry in enumerate(self._cost_log, 5):
            vals = [
                entry.get("task_id", ""),
                entry.get("product", ""),
                entry.get("character", ""),
                entry.get("reference", ""),
                entry.get("status", ""),
                entry.get("elapsed_seconds", 0),
                entry.get("estimated_cost", 0),
                entry.get("timestamp", ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # Summary row
        total_row = len(self._cost_log) + 6
        ws.merge_cells(f"A{total_row}:F{total_row}")
        ws.cell(row=total_row, column=1, value="总计").font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=7, value=round(sum(e.get("estimated_cost", 0) for e in self._cost_log), 2))
        ws.cell(row=total_row, column=7).font = Font(bold=True, color="e94560", size=12)

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 22

        wb.save(path)
        logger.info(f"XLSX 成本报告: {path}")
